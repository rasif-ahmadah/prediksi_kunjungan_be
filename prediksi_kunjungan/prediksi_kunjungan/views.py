from django.http import HttpResponse
from django.shortcuts import render
import json
from http import HTTPStatus
import os
from django.core.files.storage import FileSystemStorage
from django.shortcuts import redirect
import base64
import socket

# library for linear regression and visualization
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from mpl_toolkits.mplot3d import Axes3D
from sklearn.linear_model import LinearRegression
import openpyxl
from openpyxl import load_workbook

def Data(data_file):
    # read data
    df = pd.read_excel(data_file)

    # fitur data / variabel
    Data.fitur = df[['Bulan', 'Tahun']]
    Data.kunjungan = df.Kunjungan

# json
def Json_decode(result):
    data = result
    data = json.dumps(data)
    Json_decode.json = data

def Modelling(jumlah_bulan, tahun):
    # modelling
    model = LinearRegression()
    model.fit(Data.fitur, Data.kunjungan)

    result = []
    # new data instance
    for i in range(jumlah_bulan):
        i += 1
        input = [[i, tahun]]
    
        # make a prediction
        kunjungan_new = np.around(model.predict(input), decimals =1)
        kunjungan_new = np.expand_dims(kunjungan_new, axis=0)
        y = np.hstack((input, kunjungan_new))
        result.append(y)
   
        i += 1
    return result

def Run(data_file, jumlah_bulan, tahun):
    Data(data_file)
    jumlah_bulan = int(jumlah_bulan)
    tahun = int(tahun)

    prediction = np.array(Modelling(jumlah_bulan, tahun))

    # show the inputs and predicted outputs
    result = []
    for i in range(len(prediction)):
        myFileName=data_file
        # load the workbook, and put the sheet into a variable
        wb = load_workbook(filename=myFileName)
        ws = wb['Sheet1']

        # max_row is a sheet function that gets the last row in a sheet.
        newRowLocation = ws.max_row +1

        # write to the cell, specifying row and column, and value.
        ws.cell(column=1,row=newRowLocation, value=prediction[i][0][0])
        ws.cell(column=2,row=newRowLocation, value=prediction[i][0][1])
        ws.cell(column=3,row=newRowLocation, value=prediction[i][0][2])
        wb.save(filename=myFileName)
        wb.close()

        data = {
            'bulan' : str(prediction[i][0][0]),
            'tahun' : str(prediction[i][0][1]),
            'hasil' : str(prediction[i][0][2])
        }
        result.append(data)
    print(result)

    # grafik
    #dfnew = pd.read_excel(myFileName)
    #graph = sns.lmplot(x="Bulan", y="Kunjungan", hue="Tahun", data=dfnew, palette="Set1")
    #graph.savefig("grafik_output.png")

    #with open('grafik_output.png', mode='rb') as file:
    #   img = file.read()
    #data['img'] = base64.encodebytes(img).decode('utf-8')

    Json_decode(result)
    return(Json_decode.json)

def index(request):
    print(request)
    if request.method == "POST" and 'data_file' in request.FILES:
        data_file = request.FILES['data_file']
        jumlah_bulan = request.POST['jumlah_bulan']
        tahun = request.POST['tahun']
        return HttpResponse(Run(data_file, jumlah_bulan, tahun))
    return HttpResponse()
    
    
